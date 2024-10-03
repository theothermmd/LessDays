# Driver version : 0.1

import openpyxl


class excel_driver() :

    def __init__(self , path : str) -> None:
        self.path : str = path

    def export_data(self) -> list :

        wb_obj = openpyxl.load_workbook(self.path).active
        corses : list = []

        for i in range(1 , wb_obj.max_row + 1) :

                corses.append(
                    {'name' : wb_obj.cell(row=i, column=1).value ,
                    'teachername' : wb_obj.cell(row=i, column=2).value,
                    'day' : wb_obj.cell(row=i, column=3).value,
                    'start' : f"{wb_obj.cell(row=i, column=4).value.hour}:{wb_obj.cell(row=i, column=4).value.second}",
                    'end' : f"{wb_obj.cell(row=i, column=5).value.hour}:{wb_obj.cell(row=i, column=5).value.second}"
                    })

        return corses
        


            